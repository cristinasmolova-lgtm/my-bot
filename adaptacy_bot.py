import logging
import os
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, \
    CallbackQueryHandler
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

# Отладка: вывод информации о файле .env
print("Текущая директория:", os.getcwd())
print("Файлы в директории:", os.listdir('.'))
print("Файл .env существует:", os.path.exists('.env'))

# Отладка: чтение содержимого .env файла
if os.path.exists('.env'):
    with open('.env', 'r', encoding='utf-8') as f:
        content = f.read()
        print("Содержимое .env файла:")
        print(repr(content))  # Покажет все символы, включая скрытые
        print("Содержимое .env файла (обычный вид):")
        print(content)

# Загружаем переменные из .env файла
try:
    load_dotenv()
    print("Загрузка .env файла выполнена успешно")
except Exception as e:
    print(f"Ошибка при загрузке .env файла: {e}")

# Получаем токен бота из .env файла
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')

# Отладка: проверка токена
print("TELEGRAM_BOT_TOKEN из .env:", TELEGRAM_BOT_TOKEN)
if TELEGRAM_BOT_TOKEN:
    print("Длина токена:", len(TELEGRAM_BOT_TOKEN))
    print("Первые 10 символов токена:", TELEGRAM_BOT_TOKEN[:10])
else:
    print("Токен не найден в .env файле")
    print("Возможные причины:")
    print("1. В .env файле нет строки TELEGRAM_BOT_TOKEN=...")
    print("2. В строке есть лишние пробелы")
    print("3. Файл сохранен в неправильной кодировке")
    print("4. Токен указан в неправильном формате")

# Дополнительная проверка
if not TELEGRAM_BOT_TOKEN:
    print("Файл .env не найден или не содержит TELEGRAM_BOT_TOKEN")
    print("Создайте файл .env с содержимым: TELEGRAM_BOT_TOKEN=8463773957:AAEWJzszm_6wmaoadAYZVArFrVvp6V5UiyU")
    exit(1)

if TELEGRAM_BOT_TOKEN == "8463773957:AAEYuHLfvOnTwidKFwIwFrWEoocEVhyKoKE":
    print("ПРЕДУПРЕЖДЕНИЕ: Вы используете скомпрометированный токен!")
    print("Рекомендуется сменить токен через @BotFather как можно скорее.")

# Пути к файлам
EXCEL_FILE_PATH = 'data/users.xlsx'
PRESENTATION_PATH = r'C:\Users\Леново\PycharmProjects\PythonProject\Добро+пожаловать+в+Сбер_короткая_compressed.pdf'
VIDEO_PATH = r'C:\Users\Леново\PycharmProjects\PythonProject\VID_20250907_122717.mp4'
CORPORATE_IMAGE_PATH = r'C:\Users\Леново\PycharmProjects\PythonProject\календарь.jpg'  # Исправлено на .jpg

# Отладка: проверка существования файлов
print(f"Файл презентации существует: {os.path.exists(PRESENTATION_PATH)}")
print(f"Файл видео существует: {os.path.exists(VIDEO_PATH)}")
print(f"Файл изображения существует: {os.path.exists(CORPORATE_IMAGE_PATH)}")

# Создаем директории если их нет
os.makedirs('data', exist_ok=True)


# Проверка существования файлов
def check_files():
    files_to_check = [
        (PRESENTATION_PATH, "презентация"),
        (VIDEO_PATH, "видео"),
        (CORPORATE_IMAGE_PATH, "изображение корпоративной культуры")
    ]

    missing_files = []
    for file_path, file_name in files_to_check:
        if not os.path.exists(file_path):
            missing_files.append(f"{file_name} ({file_path})")

    if missing_files:
        print("ВНИМАНИЕ: Не найдены следующие файлы:")
        for file_desc in missing_files:
            print(f"  - {file_desc}")
        print("Пожалуйста, убедитесь, что файлы находятся по указанным путям.")
        return False
    return True


# Этапы разговора
NAME, TABEL_NUMBER, FIRST_DAY = range(3)

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# Проверка существования Excel файла и создание при необходимости
def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Дата и время", "Имя пользователя ТГ", "Имя пользователя",
                   "Табельный номер", "Дата первого дня", "User ID"])
        wb.save(EXCEL_FILE_PATH)
        print(f"Файл Excel создан: {EXCEL_FILE_PATH}")
    else:
        print(f"Файл Excel уже существует: {EXCEL_FILE_PATH}")


# Сохранение данных пользователя в Excel
def save_user_data(update, user_name, tabel_number, first_day):
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb["Users"]

        # Добавляем новую строку с данными
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            update.effective_user.username or "Не указан",
            user_name,
            tabel_number,
            first_day,
            update.effective_user.id
        ]
        ws.append(row)
        wb.save(EXCEL_FILE_PATH)
        logger.info(f"Данные пользователя {user_name} сохранены в Excel")
        print(f"Данные пользователя {user_name} сохранены в Excel")
    except Exception as e:
        logger.error(f"Ошибка при сохранении данных в Excel: {e}")
        print(f"Ошибка при сохранении данных в Excel: {e}")


# Обработчик команды /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"Пользователь {update.effective_user.first_name} начал общение")
    await update.message.reply_text(
        'Давай познакомимся! Как я могу к тебе обращаться? (твое имя)'
    )
    return NAME


# Обработчик ввода имени
async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.message.text
    context.user_data['name'] = user_name
    print(f"Получено имя: {user_name}")
    await update.message.reply_text('Напиши свой табельный номер, чтобы я смог тебя идентифицировать')
    return TABEL_NUMBER


# Обработчик ввода табельного номера
async def get_tabel_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tabel_number = update.message.text
    context.user_data['tabel_number'] = tabel_number
    print(f"Получен табельный номер: {tabel_number}")
    await update.message.reply_text('Напиши дату своего первого рабочего дня в формате дд.мм.гггг')
    return FIRST_DAY


# Обработчик ввода даты первого рабочего дня
async def get_first_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    first_day = update.message.text
    context.user_data['first_day'] = first_day
    print(f"Получена дата первого дня: {first_day}")

    # Сохраняем данные в Excel
    save_user_data(update, context.user_data['name'],
                   context.user_data['tabel_number'], first_day)

    # Создаем клавиатуру для меню
    keyboard = [
        [InlineKeyboardButton("Уральский код", callback_data='ural_code')],
        [InlineKeyboardButton("Приветствие топ-менеджеров", callback_data='greeting')],
        [InlineKeyboardButton("Корпоративная культура", callback_data='culture')],
        [InlineKeyboardButton("Контакты", callback_data='contacts')],
        [InlineKeyboardButton("Моя адаптация", callback_data='adaptation')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        'Я очень рад знакомству! Выбери пункт меню:',
        reply_markup=reply_markup
    )
    return ConversationHandler.END


# Обработчик кнопки "Назад в меню"
async def back_to_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Создаем клавиатуру для меню
    keyboard = [
        [InlineKeyboardButton("Уральский код", callback_data='ural_code')],
        [InlineKeyboardButton("Приветствие топ-менеджеров", callback_data='greeting')],
        [InlineKeyboardButton("Корпоративная культура", callback_data='culture')],
        [InlineKeyboardButton("Контакты", callback_data='contacts')],
        [InlineKeyboardButton("Моя адаптация", callback_data='adaptation')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        'Выбери пункт меню:',
        reply_markup=reply_markup
    )


# Обработчик кнопки "Уральский код"
async def send_ural_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    await query.message.reply_text(
        "Самую важную информацию про Сбер в целом и Уральский банк в частности я собрал для тебя в презентации - изучи ее пожалуйста")

    # Отправляем PDF файл
    if os.path.exists(PRESENTATION_PATH):
        try:
            with open(PRESENTATION_PATH, 'rb') as pdf_file:
                await query.message.reply_document(document=pdf_file, filename="Добро_пожаловать_в_Сбер.pdf")
            print("Презентация отправлена успешно")
        except Exception as e:
            await query.message.reply_text(f"Произошла ошибка при отправке файла: {e}")
            logger.error(f"Ошибка при отправке презентации: {e}")
            print(f"Ошибка при отправке презентации: {e}")
    else:
        await query.message.reply_text(f"Файл презентации не найден по пути: {PRESENTATION_PATH}")
        print(f"Файл презентации не найден: {PRESENTATION_PATH}")


# Обработчик кнопки "Приветствие топ-менеджеров"
async def send_greeting(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    await query.message.reply_text(
        "Сбер – это большая семья! Тебя приветствуют наши топ-менеджеры. Ознакомься, в чем заключается философия командных целей Уральского банка")

    # Отправляем видео файл
    if os.path.exists(VIDEO_PATH):
        try:
            with open(VIDEO_PATH, 'rb') as video_file:
                await query.message.reply_video(video=video_file)
            print("Видео отправлено успешно")
        except Exception as e:
            await query.message.reply_text(f"Произошла ошибка при отправке видео: {e}")
            logger.error(f"Ошибка при отправке видео: {e}")
            print(f"Ошибка при отправке видео: {e}")
    else:
        await query.message.reply_text(f"Видео файл не найден по пути: {VIDEO_PATH}")
        print(f"Файл видео не найден: {VIDEO_PATH}")


# Обработчик кнопки "Корпоративная культура"
async def send_corporate_culture(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    text = ("Сбер – это больше, чем работа! Пой, танцуй, развивай таланты вместе со Сбером!\n"
            "• Команда КВН\n"
            "• Различные мастер-классы, «Добрая ярмарка»\n"
            "• Корпоративные мероприятия, сплавы, экскурсии\n"
            "• Музыкальная кавер-группа SberBand\n"
            "• Концерты классической музыки\n"
            "• Волонтерство")

    await query.message.reply_text(text)

    # Отправляем изображение
    if os.path.exists(CORPORATE_IMAGE_PATH):
        try:
            with open(CORPORATE_IMAGE_PATH, 'rb') as image_file:
                await query.message.reply_photo(photo=image_file)
            print("Изображение корпоративной культуры отправлено успешно")
        except Exception as e:
            await query.message.reply_text(f"Произошла ошибка при отправке изображения: {e}")
            logger.error(f"Ошибка при отправке изображения: {e}")
            print(f"Ошибка при отправке изображения: {e}")
    else:
        await query.message.reply_text(f"Изображение не найдено по пути: {CORPORATE_IMAGE_PATH}")
        print(f"Файл изображения не найден: {CORPORATE_IMAGE_PATH}")

    # Клавиатура для выбора месяца
    keyboard = [
        [InlineKeyboardButton("Октябрь", callback_data='october')],
        [InlineKeyboardButton("Ноябрь", callback_data='november')],
        [InlineKeyboardButton("Декабрь", callback_data='december')],
        [InlineKeyboardButton("Назад в меню", callback_data='back_to_menu')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text('Будь в курсе мероприятий и событий Уральского банка. Выбери месяц:',
                                   reply_markup=reply_markup)


# Обработчик выбора месяца
async def send_calendar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    month = query.data

    # Отправляем календарь для октября
    if month == 'october':
        if os.path.exists(CORPORATE_IMAGE_PATH):  # Используем тот же файл для календаря
            try:
                with open(CORPORATE_IMAGE_PATH, 'rb') as image_file:
                    await query.message.reply_photo(photo=image_file)
                print("Календарь отправлен успешно")
            except Exception as e:
                await query.message.reply_text(f"Произошла ошибка при отправке изображения: {e}")
                logger.error(f"Ошибка при отправке календаря: {e}")
                print(f"Ошибка при отправке календаря: {e}")
        else:
            await query.message.reply_text(f"Календарь на октябрь не найден по пути: {CORPORATE_IMAGE_PATH}")
            print(f"Файл календаря не найден: {CORPORATE_IMAGE_PATH}")

    # Клавиатура с кнопкой "Назад"
    keyboard = [
        [InlineKeyboardButton("Назад в меню", callback_data='back_to_menu')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text('Выберите другое действие:', reply_markup=reply_markup)


# Обработчик кнопки "Контакты"
async def send_contacts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    await query.message.reply_text(
        "Любые вопросы направляй на почту куратора по адаптации в Уральском банке Котельниковой Кристине Kotelnikova.K.A@sberbank.ru")


# Обработчик кнопки "Моя адаптация"
async def send_adaptation_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    text = ("Поздравляю! Ты узнал много нового про Сбер, Урал и возможности сотрудников! "
            "Теперь предлагаю тебе погрузиться в процесс именно твоей адаптации - "
            "я расскажу тебе про особенности адаптации в Сбере и напомню о важных шагах "
            "именно твоей адаптации и дальнейшего развития. "
            "Проверь себя по чек-листу новичка, какие шаги у тебя пройдены на текущий момент.")

    await query.message.reply_text(text)

    # Отправляем чек-лист
    checklist_text = ("Поставь галочку (нажми на номер) напротив шагов, которые ты уже прошел:\n\n"
                      "1. у меня есть рабочее место, оборудование и получены доступы\n"
                      "2. меня представили команде и познакомили с коллегами\n"
                      "3. я знаю своего бадди, его ФИО есть в Пульс\n"
                      "4. я заполнил свой профиль в Пульс\n"
                      "5. я начал проходить трек адаптации в Пульс\n"
                      "6. руководитель провел со мной 1:1 встречу\n"
                      "7. мне понятны мои цели на испытательный срок, я внес их в Пульс\n"
                      "8. я вступил в чат новичков в Сберчате\n"
                      "9. в Пульс мне приходят опросы удовлетворенности процессом адаптации, я оставляю обратную связь по процессу\n"
                      "10. к концу испытательного срока я завершил все назначенные обязательные обучения\n"
                      "11. руководитель провел со мной 1:1 встречу по подведению итогов адаптации\n"
                      "12. мне понятен итог периода моей адаптации")

    # Клавиатура для чек-листа
    keyboard = [
        [InlineKeyboardButton(str(i), callback_data=f'check_{i}') for i in range(1, 7)],
        [InlineKeyboardButton(str(i), callback_data=f'check_{i}') for i in range(7, 13)],
        [InlineKeyboardButton("Завершить чек-лист", callback_data='finish_checklist')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text(checklist_text, reply_markup=reply_markup)


# Обработчик выбора элемента чек-листа
async def toggle_checklist_item(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    item_number = query.data.split('_')[1]

    # Инициализируем список выполненных пунктов
    if 'completed_items' not in context.user_data:
        context.user_data['completed_items'] = set()

    # Переключаем статус выполнения
    if item_number in context.user_data['completed_items']:
        context.user_data['completed_items'].remove(item_number)
        # Обновляем сообщение с чек-листом
        checklist_text = ("Поставь галочку (нажми на номер) напротив шагов, которые ты уже прошел:\n\n"
                          "1. у меня есть рабочее место, оборудование и получены доступы\n"
                          "2. меня представили команде и познакомили с коллегами\n"
                          "3. я знаю своего бадди, его ФИО есть в Пульс\n"
                          "4. я заполнил свой профиль в Пульс\n"
                          "5. я начал проходить трек адаптации в Пульс\n"
                          "6. руководитель провел со мной 1:1 встречу\n"
                          "7. мне понятны мои цели на испытательный срок, я внес их в Пульс\n"
                          "8. я вступил в чат новичков в Сберчате\n"
                          "9. в Пульс мне приходят опросы удовлетворенности процессом адаптации, я оставляю обратную связь по процессу\n"
                          "10. к концу испытательного срока я завершил все назначенные обязательные обучения\n"
                          "11. руководитель провел со мной 1:1 встречу по подведению итогов адаптации\n"
                          "12. мне понятен итог периода моей адаптации")

        # Обновляем клавиатуру с учетом отмеченных элементов
        keyboard = []
        for i in range(1, 7):
            if str(i) in context.user_data['completed_items']:
                keyboard.append(InlineKeyboardButton(f"✓ {i}", callback_data=f'check_{i}'))
            else:
                keyboard.append(InlineKeyboardButton(str(i), callback_data=f'check_{i}'))

        keyboard_row1 = [keyboard[0], keyboard[1], keyboard[2], keyboard[3], keyboard[4], keyboard[5]]

        keyboard2 = []
        for i in range(7, 13):
            if str(i) in context.user_data['completed_items']:
                keyboard2.append(InlineKeyboardButton(f"✓ {i}", callback_data=f'check_{i}'))
            else:
                keyboard2.append(InlineKeyboardButton(str(i), callback_data=f'check_{i}'))

        keyboard_row2 = [keyboard2[0], keyboard2[1], keyboard2[2], keyboard2[3], keyboard2[4], keyboard2[5]]

        keyboard_final = [keyboard_row1, keyboard_row2,
                          [InlineKeyboardButton("Завершить чек-лист", callback_data='finish_checklist')]]
        reply_markup = InlineKeyboardMarkup(keyboard_final)

        await query.message.edit_reply_markup(reply_markup=reply_markup)
        print(f"Пункт {item_number} снят с чек-листа")
    else:
        context.user_data['completed_items'].add(item_number)
        # Обновляем сообщение с чек-листом
        checklist_text = ("Поставь галочку (нажми на номер) напротив шагов, которые ты уже прошел:\n\n"
                          "1. у меня есть рабочее место, оборудование и получены доступы\n"
                          "2. меня представили команде и познакомили с коллегами\n"
                          "3. я знаю своего бадди, его ФИО есть в Пульс\n"
                          "4. я заполнил свой профиль в Пульс\n"
                          "5. я начал проходить трек адаптации в Пульс\n"
                          "6. руководитель провел со мной 1:1 встречу\n"
                          "7. мне понятны мои цели на испытательный срок, я внес их в Пульс\n"
                          "8. я вступил в чат новичков в Сберчате\n"
                          "9. в Пульс мне приходят опросы удовлетворенности процессом адаптации, я оставляю обратную связь по процессу\n"
                          "10. к концу испытательного срока я завершил все назначенные обязательные обучения\n"
                          "11. руководитель провел со мной 1:1 встречу по подведению итогов адаптации\n"
                          "12. мне понятен итог периода моей адаптации")

        # Обновляем клавиатуру с учетом отмеченных элементов
        keyboard = []
        for i in range(1, 7):
            if str(i) in context.user_data['completed_items']:
                keyboard.append(InlineKeyboardButton(f"✓ {i}", callback_data=f'check_{i}'))
            else:
                keyboard.append(InlineKeyboardButton(str(i), callback_data=f'check_{i}'))

        keyboard_row1 = [keyboard[0], keyboard[1], keyboard[2], keyboard[3], keyboard[4], keyboard[5]]

        keyboard2 = []
        for i in range(7, 13):
            if str(i) in context.user_data['completed_items']:
                keyboard2.append(InlineKeyboardButton(f"✓ {i}", callback_data=f'check_{i}'))
            else:
                keyboard2.append(InlineKeyboardButton(str(i), callback_data=f'check_{i}'))

        keyboard_row2 = [keyboard2[0], keyboard2[1], keyboard2[2], keyboard2[3], keyboard2[4], keyboard2[5]]

        keyboard_final = [keyboard_row1, keyboard_row2,
                          [InlineKeyboardButton("Завершить чек-лист", callback_data='finish_checklist')]]
        reply_markup = InlineKeyboardMarkup(keyboard_final)

        await query.message.edit_reply_markup(reply_markup=reply_markup)
        print(f"Пункт {item_number} добавлен в чек-лист")


# Обработчик завершения чек-листа
async def finish_checklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    completed_count = len(context.user_data.get('completed_items', set()))
    total_count = 12

    await query.message.reply_text(
        f"Ты отметил {completed_count} из {total_count} шагов чек-листа.\n\n"
        "Данные чек-лист поможет тебе держать в фокусе основные шаги адаптации. "
        "А если у тебя возникли вопросы по любому из пунктов, в основном меню нажми кнопку «Контакты»."
    )

    # Возвращаемся в главное меню
    keyboard = [
        [InlineKeyboardButton("Уральский код", callback_data='ural_code')],
        [InlineKeyboardButton("Приветствие топ-менеджеров", callback_data='greeting')],
        [InlineKeyboardButton("Корпоративная культура", callback_data='culture')],
        [InlineKeyboardButton("Контакты", callback_data='contacts')],
        [InlineKeyboardButton("Моя адаптация", callback_data='adaptation')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text('Выбери пункт меню:', reply_markup=reply_markup)


# Обработчик возврата в главное меню
async def return_to_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Создаем клавиатуру для меню
    keyboard = [
        [InlineKeyboardButton("Уральский код", callback_data='ural_code')],
        [InlineKeyboardButton("Приветствие топ-менеджеров", callback_data='greeting')],
        [InlineKeyboardButton("Корпоративная культура", callback_data='culture')],
        [InlineKeyboardButton("Контакты", callback_data='contacts')],
        [InlineKeyboardButton("Моя адаптация", callback_data='adaptation')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.edit_text('Выбери пункт меню:', reply_markup=reply_markup)


def main():
    print("=== Запуск бота ===")

    # Проверяем наличие необходимых файлов
    if not check_files():
        print("Некоторые файлы отсутствуют. Бот запустится, но функциональность может быть ограничена.")

    # Создаем Excel файл если его нет
    ensure_excel_file()

    # Создаем приложение
    print("Создание приложения...")
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    print("Приложение создано успешно")

    # Создаем разговорный обработчик
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            TABEL_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_tabel_number)],
            FIRST_DAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_first_day)]
        },
        fallbacks=[]
    )

    # Добавляем обработчики команд
    application.add_handler(conv_handler)

    # Обработчики кнопок
    application.add_handler(CallbackQueryHandler(send_ural_code, pattern='^ural_code$'))
    application.add_handler(CallbackQueryHandler(send_greeting, pattern='^greeting$'))
    application.add_handler(CallbackQueryHandler(send_corporate_culture, pattern='^culture$'))
    application.add_handler(CallbackQueryHandler(send_calendar, pattern='^(october|november|december)$'))
    application.add_handler(CallbackQueryHandler(back_to_menu, pattern='^back_to_menu$'))
    application.add_handler(CallbackQueryHandler(send_contacts, pattern='^contacts$'))
    application.add_handler(CallbackQueryHandler(send_adaptation_info, pattern='^adaptation$'))
    application.add_handler(CallbackQueryHandler(toggle_checklist_item, pattern='^check_'))
    application.add_handler(CallbackQueryHandler(finish_checklist, pattern='^finish_checklist$'))
    application.add_handler(CallbackQueryHandler(return_to_main_menu, pattern='^main_menu$'))

    print("Бот запускается...")
    # Запускаем бота
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()