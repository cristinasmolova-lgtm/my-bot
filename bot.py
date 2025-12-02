# bot.py
import asyncio
import logging
from datetime import datetime
from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.filters import CommandStart
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook, load_workbook
from config import BOT_TOKEN, EXCEL_FILE_PATH, PDF_PATH_1, PDF_PATH_2, P2P_IMAGE_PATH, EVENT_IMAGE_PATH, NEWS_IMAGE_PATH_1, NEWS_IMAGE_PATH_2, YANDEX_DISK_URL

# Настройка логирования
logging.basicConfig(level=logging.INFO)

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# FSM States
class RegistrationStates(StatesGroup):
    waiting_for_name = State()
    waiting_for_employee_id = State()
    waiting_for_start_date = State()

class FeedbackStates(StatesGroup):
    waiting_for_feedback_1 = State()
    waiting_for_feedback_2 = State()
    waiting_for_feedback_3 = State()

class QuestionStates(StatesGroup):
    waiting_for_question = State()

# Глобальные переменные для хранения данных пользователя
user_data = {}

# Функция для создания или загрузки Excel файла
def ensure_excel_file():
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        # Заголовки
        headers = ["Дата и время", "Username Telegram", "Имя пользователя", "Табельный номер", "Дата первого дня", "User ID", "Доп. информация"]
        ws.append(headers)
        wb.save(EXCEL_FILE_PATH)
    return wb

# Функция для записи данных в Excel
def write_to_excel(user_id, username, user_input, additional_info=""):
    wb = ensure_excel_file()
    ws = wb.active
    row = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        username,
        user_data.get(user_id, {}).get('name', ''),
        user_data.get(user_id, {}).get('employee_id', ''),
        user_data.get(user_id, {}).get('start_date', ''),
        user_id,
        additional_info or user_input
    ]
    ws.append(row)
    wb.save(EXCEL_FILE_PATH)

# Команда /start
@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    user_id = message.from_user.id
    username = message.from_user.username or "N/A"

    # Инициализируем данные пользователя, если их нет
    if user_id not in user_data:
        user_data[user_id] = {'username': username}

    await state.set_state(RegistrationStates.waiting_for_name)
    await message.answer("Давай знакомиться! Напиши свое имя")

# Обработчик имени
@dp.message(RegistrationStates.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['name'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Имя пользователя")

    await state.set_state(RegistrationStates.waiting_for_employee_id)
    await message.answer("Напиши свой табельный номер, чтобы я мог найти в системе")

# Обработчик табельного номера
@dp.message(RegistrationStates.waiting_for_employee_id)
async def process_employee_id(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['employee_id'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Табельный номер")

    await state.set_state(RegistrationStates.waiting_for_start_date)
    await message.answer("Напиши дату своего первого рабочего дня, чтобы мы могли присылать тебе уведомления и важные напоминания")

# Обработчик даты первого дня
@dp.message(RegistrationStates.waiting_for_start_date)
async def process_start_date(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['start_date'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Дата первого рабочего дня")

    await state.clear() # Завершаем FSM регистрация
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="1. Сбер на Урале", callback_data="info_sber")],
        [InlineKeyboardButton(text="2. Видео", callback_data="info_video")],
        [InlineKeyboardButton(text="3. Peer-to-peer", callback_data="info_p2p")],
        [InlineKeyboardButton(text="4. Культура и сообщества", callback_data="info_culture")],
        [InlineKeyboardButton(text="5. Это все мое", callback_data="info_benefits")],
        [InlineKeyboardButton(text="6. Контакты", callback_data="info_contacts")],
        [InlineKeyboardButton(text="7. Оставить обратную связь", callback_data="feedback_start")],
        [InlineKeyboardButton(text="8. Новости", callback_data="info_news")],
        [InlineKeyboardButton(text="Задать вопрос", callback_data="ask_question")],
    ])
    await message.answer("Рад знакомству! Выбери пункт меню и изучай материалы:", reply_markup=keyboard)

# Обработчик кнопок меню
@dp.callback_query(F.data == "info_sber")
async def send_info_sber(callback_query: CallbackQuery):
    await callback_query.answer() # Ответ на callback
    await callback_query.message.answer("Самую важную информацию про Сбер и Урал я собрал для тебя в презентации - изучай, задавай вопросы, если есть")
    try:
        document = FSInputFile(PDF_PATH_1)
        await callback_query.message.answer_document(document)
    except Exception as e:
        await callback_query.message.answer(f"Не удалось отправить файл: {e}")
        logging.error(f"Error sending PDF: {e}")

@dp.callback_query(F.data == "info_video")
async def send_info_video(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer(f"Ты стал частью большой команды Сбера и тебя приветствуют наши топ-менеджеры. Смотри видео.\n{YANDEX_DISK_URL}")

@dp.callback_query(F.data == "info_p2p")
async def send_info_p2p(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer("На всем периоде адаптации твоя основная поддержка - это HR-платформа Пульс и твой бадди. Не забывай просматривать уведомления и задачи, проходи индивидуальный трек адаптации. А бадди - это один из представителей ролей взаимного развития (peеr-to-peеr). Культура взаимного развития - это также консультанты по развитию, коучи, наставники, фасилитаторы, медиаторы. Подробнее ты сможешь ознакомиться в Пульс (раздел Развитие)")
    try:
        image = FSInputFile(P2P_IMAGE_PATH)
        await callback_query.message.answer_photo(image)
    except Exception as e:
        await callback_query.message.answer(f"Не удалось отправить изображение: {e}")
        logging.error(f"Error sending P2P image: {e}")

@dp.callback_query(F.data == "info_culture")
async def send_info_culture(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer("Уральский банк живет насыщенной культурной и спортивной жизнью. Обязательно присоединяйся к мероприятиям - вся информация приходит тебе на почту. Вот несколько фото с последних событий")
    try:
        image = FSInputFile(EVENT_IMAGE_PATH)
        await callback_query.message.answer_photo(image)
    except Exception as e:
        await callback_query.message.answer(f"Не удалось отправить изображение: {e}")
        logging.error(f"Error sending event image: {e}")

    await callback_query.message.answer("Вступай в сообщества Уральского банка - будь в курсе событий!\n🗣️ Телеграм-канал \"Говорит Урал\" — новости, анонсы, важные события\n🤝 Телеграм-канал \"Биржа волонтёров Екатеринбург (УБ)\" — анонсы, поддержка, активности Ссылки на каналы находятся в презентации, которую ты изучил выше. Вопросы? Пиши в раздел «Контакты»!")

@dp.callback_query(F.data == "info_benefits")
async def send_info_benefits(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer("Сбер заботится о свои сотрудниках с самого первого дня работы. В презентации собрали для вас все корпоративные льготы и привилегии. Изучай, пользуйся - ведь это все твое!")
    try:
        document = FSInputFile(PDF_PATH_2)
        await callback_query.message.answer_document(document)
    except Exception as e:
        await callback_query.message.answer(f"Не удалось отправить файл: {e}")
        logging.error(f"Error sending benefits PDF: {e}")

@dp.callback_query(F.data == "info_contacts")
async def send_info_contacts(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer("Любые вопросы направляй на почту куратора по адаптации в Уральском банке Котельниковой Кристине Kotelnikova.K.A@sberbank.ru")

@dp.callback_query(F.data == "info_news")
async def send_info_news(callback_query: CallbackQuery):
    await callback_query.answer()
    await callback_query.message.answer("22 октября в Технохабе Екатеринбурга прошла встреча Вице-президента-председателя Колтыпина Петра Николаевича и Заместителя председателя, руководителя блока Люди и культура Осиповой Марии Леонидовны с новыми сотрудниками команды Сбера на Урале. На встрече обсудили особенности бизнеса на Урале, какими качествами и ценностями должны обладать сотрудники Сбера и как достигать карьерных высот. Такие мероприятия заряжают энергией и успехом!")
    try:
        image1 = FSInputFile(NEWS_IMAGE_PATH_1)
        image2 = FSInputFile(NEWS_IMAGE_PATH_2)
        await callback_query.message.answer_photo(image1)
        await callback_query.message.answer_photo(image2)
    except Exception as e:
        await callback_query.message.answer(f"Не удалось отправить изображения: {e}")
        logging.error(f"Error sending news images: {e}")

# Обработчик начала обратной связи
@dp.callback_query(F.data == "feedback_start")
async def start_feedback(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.answer()
    await state.set_state(FeedbackStates.waiting_for_feedback_1)
    await callback_query.message.answer("1. Опиши, что понравилось при использовании бота")

# Обработчики обратной связи
@dp.message(FeedbackStates.waiting_for_feedback_1)
async def feedback_1(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['feedback_1'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Обратная связь 1")
    await state.set_state(FeedbackStates.waiting_for_feedback_2)
    await message.answer("2. Напиши, чего тебе не хватило при использовании бота")

@dp.message(FeedbackStates.waiting_for_feedback_2)
async def feedback_2(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['feedback_2'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Обратная связь 2")
    await state.set_state(FeedbackStates.waiting_for_feedback_3)
    await message.answer("3. Что можно добавить в чат-бот, чтобы его использование было максимально полезным для новых сотрудников")

@dp.message(FeedbackStates.waiting_for_feedback_3)
async def feedback_3(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['feedback_3'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Обратная связь 3")
    await state.clear()
    await message.answer("Благодарим за обратную связь, это очень важно для дальнейшего развития нашего виртуального помощника. Среди всех участников опроса первого числа каждого календарного месяца мы будем разыгрывать памятный мерч - следи за уведомлениями!")

# Обработчик кнопки "Задать вопрос"
@dp.callback_query(F.data == "ask_question")
async def ask_question_prompt(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.answer()
    await state.set_state(QuestionStates.waiting_for_question)
    await callback_query.message.answer("Напиши свой вопрос")

# Обработчик вопроса
@dp.message(QuestionStates.waiting_for_question)
async def receive_question(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_data[user_id]['question'] = message.text
    write_to_excel(user_id, message.from_user.username or "N/A", message.text, "Вопрос пользователя")
    await state.clear()
    await message.answer("Благодарим за твой вопрос. Взяли в работу - вернемся с ответом")

# Запуск бота
if __name__ == '__main__':
    ensure_excel_file() # Убедимся, что Excel файл создан
    asyncio.run(dp.start_polling(bot))
